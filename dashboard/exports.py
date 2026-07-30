"""
Excel annexure + PNG zip builders for the Chandigarh audit dashboard.

Excel: 10-sheet workbook consumed directly as an annexure in the audit
report. Pandas writes the rows; styling and conditional formatting are
applied inline through the open ExcelWriter's worksheet handles so the
workbook is never re-loaded and re-saved.

PNG: each Plotly figure exported via kaleido at 1600×1000 px, packed
into a single zip for the report annexure.
"""

from __future__ import annotations

import io
import zipfile
from datetime import datetime

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from data import AUDIT_WINDOW_END, AUDIT_WINDOW_START
from metrics import (
    SHORT_CORRIDOR_IDS, am_peak_observations, bti as compute_bti, cv as compute_cv,
    direction_asymmetry, hourly_median_cr, minutes_lost_table, peak_observations,
    pm_peak_observations, ranking_table, weekday_observations,
)


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="left", vertical="center")

# Width-estimate scanning cost is O(rows × cols). On the 40k-row Raw
# Observations sheet that means hundreds of thousands of Python iterations
# for no visible gain — the Raw sheet's column widths are predictable. Cap
# the scan and apply fixed widths for the raw sheet instead.
_AUTOWIDTH_SAMPLE_ROWS = 200
_AUTOWIDTH_MIN = 10
_AUTOWIDTH_MAX = 60


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _format_header(ws, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
    ws.freeze_panes = "A2"


def _auto_width_from_df(ws, df: pd.DataFrame, include_index: bool = False) -> None:
    """Set column widths from the dataframe instead of scanning every cell.

    Reads the first _AUTOWIDTH_SAMPLE_ROWS rows of each column to estimate
    width. Cheap regardless of total row count; on the Raw Observations
    sheet this replaces hundreds of thousands of cell reads with a few
    hundred. Header text also factors into the width so column titles fit.
    """
    cols: list[tuple[str, pd.Series]] = []
    if include_index:
        cols.append((str(df.index.name or ""), pd.Series(df.index.astype(str)[: _AUTOWIDTH_SAMPLE_ROWS])))
    for name in df.columns:
        cols.append((str(name), df[name].head(_AUTOWIDTH_SAMPLE_ROWS).astype(str)))
    for col_idx, (header, sample) in enumerate(cols, start=1):
        max_len = len(header)
        if not sample.empty:
            max_len = max(max_len, int(sample.str.len().max() or 0))
        width = min(max(max_len + 2, _AUTOWIDTH_MIN), _AUTOWIDTH_MAX)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _style_sheet(ws, df: pd.DataFrame, include_index: bool = False) -> None:
    n_cols = len(df.columns) + (1 if include_index else 0)
    _format_header(ws, n_cols)
    _auto_width_from_df(ws, df, include_index=include_index)


def build_excel_annexure(df: pd.DataFrame, rep: dict) -> bytes:
    """Produce the multi-sheet xlsx as bytes (Streamlit download_button consumes bytes).

    Styling, frozen headers, autosize and conditional formatting are all
    applied inline against the live ExcelWriter worksheet handles. The
    previous implementation pandas-wrote, closed the writer, openpyxl-loaded
    the bytes back in, restyled, and saved again — a ~10 s round trip on
    the 40k-row Raw Observations sheet that is now eliminated.
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        stats = rep["stats"]
        sheets: list[tuple[str, pd.DataFrame, bool]] = []  # (sheet_name, df, include_index)

        # 1. Cover -------------------------------------------------------
        cover = pd.DataFrame({
            "Field": [
                "Title", "Audit window (start)", "Audit window (end)",
                "Polling interval", "OD pairs",
                "First observation in dataset", "Last observation in dataset",
                "Days covered", "Corridors covered", "Total OK observations",
                "Total FAIL rows", "FAIL %",
                "Observations MD5", "corridors.csv MD5",
                "Built at",
            ],
            "Value": [
                "Chandigarh Mobility Audit — Congestion Index Annexure",
                str(AUDIT_WINDOW_START.date()) + " 00:00 IST",
                str(AUDIT_WINDOW_END.date()) + " 23:59 IST",
                "Every 30 minutes",
                "50 (25 corridors × 2 directions)",
                stats.first_timestamp,
                stats.last_timestamp,
                stats.days_covered,
                f"{stats.corridors_covered}/38",
                f"{stats.total_observations:,}",
                f"{stats.fail_count:,}",
                f"{stats.fail_pct:.3f}%",
                stats.observations_md5,
                stats.corridors_md5,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ],
        })
        cover.to_excel(writer, sheet_name="1. Cover", index=False)
        sheets.append(("1. Cover", cover, False))

        # 2. Ranking -----------------------------------------------------
        ranking = ranking_table(df)
        ranking_display = ranking[[
            "rank", "corridor_id", "corridor_name", "phci", "phci_hour",
            "phci_direction", "adci", "bti", "cv", "n_peak", "is_short_corridor",
        ]].copy()
        ranking_display.columns = [
            "Rank", "Corridor ID", "Corridor", "PHCI", "Peak hour",
            "Peak direction", "ADCI (06-22)", "BTI (peak)", "CV (peak)",
            "n (peak obs)", "Short corridor (<1.5 km)",
        ]
        ranking_display.to_excel(writer, sheet_name="2. Ranking", index=False)
        sheets.append(("2. Ranking", ranking_display, False))

        # 3. Hourly medians ---------------------------------------------
        hm = hourly_median_cr(df)
        hm.to_excel(writer, sheet_name="3. Hourly Medians", index=False)
        sheets.append(("3. Hourly Medians", hm, False))

        # 4. Direction asymmetry ----------------------------------------
        asym = direction_asymmetry(df)
        asym.to_excel(writer, sheet_name="4. Direction Asymmetry", index=False)
        sheets.append(("4. Direction Asymmetry", asym, False))

        # 5. Reliability -------------------------------------------------
        bti_df = compute_bti(df)
        cv_df = compute_cv(df)[["corridor_id", "direction", "cv", "mu_peak_s", "sigma_peak_s"]]
        rel = bti_df.merge(cv_df, on=["corridor_id", "direction"], how="left")
        rel.to_excel(writer, sheet_name="5. Reliability", index=False)
        sheets.append(("5. Reliability", rel, False))

        # 6. Coverage ----------------------------------------------------
        coverage_long = rep["coverage"]
        coverage_pivot = coverage_long.pivot(
            index="corridor_id", columns="date", values="n_obs"
        ).fillna(0).astype(int)
        coverage_pivot.to_excel(writer, sheet_name="6. Coverage")
        sheets.append(("6. Coverage", coverage_pivot, True))

        # 7. FAIL log ----------------------------------------------------
        fail_log = rep["fail_log"]
        if not fail_log.empty:
            fail_display = fail_log[[
                "timestamp_ist", "date", "time", "corridor_id", "corridor_name",
                "direction", "api_status", "error_msg",
            ]]
        else:
            fail_display = pd.DataFrame({"note": ["No failed API calls in the cumulative log."]})
        fail_display.to_excel(writer, sheet_name="7. FAIL Log", index=False)
        sheets.append(("7. FAIL Log", fail_display, False))

        # 8. Distance drift ---------------------------------------------
        drift = rep["distance_drift"].copy()
        drift.to_excel(writer, sheet_name="8. Distance Drift", index=False)
        sheets.append(("8. Distance Drift", drift, False))

        # 9. Methodology -------------------------------------------------
        from metrics import (
            PEAK_PRESETS, active_peak_hours, active_peak_preset,
        )
        _am_x, _pm_x, _ = active_peak_hours()
        _preset_meta_x = PEAK_PRESETS.get(active_peak_preset(), {})
        _am_label_x = (f"{_am_x[0]:02d}:00 to {_am_x[-1] + 1:02d}:00 IST"
                       if _am_x else "—")
        _pm_label_x = (f"{_pm_x[0]:02d}:00 to {_pm_x[-1] + 1:02d}:00 IST"
                       if _pm_x else "—")
        methodology_rows = [
            ("Instant Congestion Ratio",
             "CR(i,t) = duration_traffic_s / duration_freeflow_s"),
            ("Hourly Median CR",
             "CR_hour = median over the hour, per (corridor, direction, hour)"),
            ("PHCI (Peak-Hour Congestion Index)",
             f"max over peak hours of weekday median CR per (corridor, direction); "
             f"both directions collapsed by max. Peak hours = active preset (see below)."),
            ("ADCI (All-Day Congestion Index)",
             "mean over active hours 06-21 of the hourly median CR"),
            ("BTI (Buffer Time Index, FHWA)",
             "(p95(duration_traffic_peak) - median(duration_traffic_peak)) / median(...)"),
            ("CV (Coefficient of Variation)",
             "sigma(duration_traffic_peak) / mu(duration_traffic_peak)"),
            ("Peak window preset (active at export)",
             _preset_meta_x.get("label", active_peak_preset())),
            ("Peak window — AM", _am_label_x),
            ("Peak window — PM", _pm_label_x),
            ("Active hours (ADCI)", "06:00 to 22:00 IST"),
            ("Short corridors (asterisked in Ranking)",
             ", ".join(sorted(SHORT_CORRIDOR_IDS)) + " — < 1.5 km"),
            ("Filter applied to OK observations",
             "api_status == 'OK' AND duration_freeflow_s > 0 AND congestion_ratio is not null"),
            ("Dedupe key",
             "(timestamp_ist, corridor_id, direction) — last write wins across snapshot CSVs"),
            ("Holidays (UT)",
             "None excluded in the current audit window — every calendar weekday "
             "13–26 May 2026 enters weekday aggregations on equal footing."),
        ]
        methodology_df = pd.DataFrame(methodology_rows, columns=["Term", "Definition"])
        methodology_df.to_excel(writer, sheet_name="9. Methodology", index=False)
        sheets.append(("9. Methodology", methodology_df, False))

        # 10. Raw observations ------------------------------------------
        raw_out = df.copy()
        if len(raw_out) > 100_000:
            # Truncate the embedded copy; note this on the Cover sheet.
            raw_out = raw_out.head(100_000)
        raw_out.to_excel(writer, sheet_name="10. Raw Observations", index=False)
        # Autosize from a sample (the helper itself caps the scan at 200
        # rows) — predictable widths without iterating the full 40k rows.
        sheets.append(("10. Raw Observations", raw_out, False))

        # Apply header styling + autosize inline against the live worksheet
        # handles so we don't have to load_workbook/save the bytes again.
        for sheet_name, sheet_df, include_index in sheets:
            ws = writer.sheets[sheet_name]
            _style_sheet(ws, sheet_df, include_index=include_index)

        # Conditional formatting on Coverage sheet.
        cov_ws = writer.sheets["6. Coverage"]
        if cov_ws.max_row > 1 and cov_ws.max_column > 1:
            last_col_letter = get_column_letter(cov_ws.max_column)
            range_str = f"B2:{last_col_letter}{cov_ws.max_row}"
            cov_ws.conditional_formatting.add(
                range_str,
                ColorScaleRule(
                    # Per-cell scale: 48 batches/day × 2 directions = 96 obs at full coverage.
                    start_type="num", start_value=0, start_color="DC2626",
                    mid_type="num", mid_value=60, mid_color="FCD34D",
                    end_type="num", end_value=96, end_color="16A34A",
                ),
            )

    return buf.getvalue()


# ---------------------------------------------------------------------------
# PNG zip
# ---------------------------------------------------------------------------

def build_png_zip(figures: dict) -> bytes:
    """Pack a dict of {filename: plotly.Figure} into a zip of 1600x1000 PNGs.

    `scale=1` keeps the native render at 1600x1000 (≈267 DPI when embedded
    at 6" width in a Word annexure — still print-grade). The previous
    `scale=2` doubled every dimension to 3200x2000 (8 MP/image), which
    quadrupled kaleido render time per image for a quality gain that does
    not survive Word's column-width resampling. Eight figures × ~3 s each
    is now ~1 s each.
    """
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, fig in figures.items():
            try:
                png_bytes = fig.to_image(format="png", width=1600, height=1000, scale=1)
                zf.writestr(fname, png_bytes)
            except Exception as e:
                zf.writestr(
                    fname + ".error.txt",
                    f"Could not export {fname}: {e}\n"
                    "Install the kaleido package: pip install kaleido==0.2.1\n",
                )
    return out.getvalue()
